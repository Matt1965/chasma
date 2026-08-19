#!/usr/bin/env py -3
"""One-shot workbook update for Cavecrawler unit integration."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Chasma Design.xlsx"
BACKUP = ROOT / "Chasma Design.xlsx.cavecrawler.bak"

ANIMATION_NEW_COLUMNS = [
    "Death Animation",
    "Hit Reaction Animation",
    "Upper Body Split Bone",
    "Turn Left Animation",
    "Turn Right Animation",
    "Turn Left Duration",
    "Turn Right Duration",
]

CAVECRAWLER_PROFILE = {
    "Profile ID": "cavecrawler",
    "Idle Animation": "Idle",
    "Walk Animation": "CrawlForward",
    "Run Animation": "",
    "Locomotion Reference Speed": 3.5,
    "Enabled": "Y",
    "Death Animation": "Death",
    "Hit Reaction Animation": "GetHit1",
    "Upper Body Split Bone": "",
    "Turn Left Animation": "CrawlLeft",
    "Turn Right Animation": "CrawlRight",
    "Turn Left Duration": 1.0,
    "Turn Right Duration": 1.0,
}

CAVECRAWLER_WEAPON = {
    "Weapon ID": "weapon_cavecrawler_claws",
    "Name": "Cavecrawler Claws",
    "Description": "Natural claw attack.",
    "Damage": 6,
    "Damage Type": "Slashing",
    "Range": 1.4,
    "Attacks Per Second": 1.25,
    "Windup": 0.18,
    "Recovery": 0.12,
    "Hit Mode": "Melee",
    "Projectile Key": None,
    "Animation Key": "ClawAttackRight",
    "Target Filters": "Enemies, Wildlife",
    "Stat Scaling": None,
    "Enabled": "Y",
}

# PLACEHOLDER catalog stats — no authoritative design source in repo.
CAVECRAWLER_UNIT = {
    "Unit ID": "U-0003",
    "Name": "Cavecrawler",
    "Faction": "Wild",
    "File Path": r"assets\units\cavecrawler.glb",
    "Level": 3,
    "Base HP": 45,
    "Strength": 8,
    "Dexterity": 6,
    "Constitution": 7,
    "Agility": 5,
    "Charisma": 2,
    "Intelligence": 3,
    "Move Speed": 3.5,
    "Collision Radius": 1.1,
    "Max Slope": 45,
    "Animation Profile": "cavecrawler",
    "Default Weapon ID": "weapon_cavecrawler_claws",
    "Enabled": "Y",
}


def header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }


def ensure_columns(ws, columns: list[str]) -> dict[str, int]:
    headers = header_map(ws)
    next_col = ws.max_column + 1
    for column in columns:
        if column not in headers:
            ws.cell(1, next_col, column)
            headers[column] = next_col
            next_col += 1
    return headers


def total_stats(row: dict[str, object]) -> float:
    return sum(
        float(row[name])
        for name in (
            "Strength",
            "Dexterity",
            "Constitution",
            "Agility",
            "Charisma",
            "Intelligence",
        )
    )


def power_rating(level: float, base_hp: float, stats: float) -> float:
    return round(level * 2 + base_hp * 0.5 + stats * 0.8, 1)


def tier_label(rating: float) -> str:
    if rating >= 25:
        return "Elite"
    if rating >= 15:
        return "Veteran"
    if rating >= 8:
        return "Regular"
    return "Rookie"


def upsert_row_by_key(ws, headers: dict[str, int], key_column: str, key_value: str, data: dict) -> None:
    key_col = headers[key_column]
    target_row = None
    for row in range(2, ws.max_row + 2):
        existing = ws.cell(row, key_col).value
        if existing is not None and str(existing).strip() == key_value:
            target_row = row
            break
        if existing is None or str(existing).strip() == "":
            target_row = row
            break
    assert target_row is not None
    for column, value in data.items():
        if column not in headers:
            continue
        ws.cell(target_row, headers[column], value)


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    shutil.copy2(WORKBOOK, BACKUP)
    wb = openpyxl.load_workbook(WORKBOOK)

    # Animation Profiles
    anim = wb["Animation Profiles"]
    anim_headers = ensure_columns(anim, ANIMATION_NEW_COLUMNS)
    upsert_row_by_key(anim, anim_headers, "Profile ID", "cavecrawler", CAVECRAWLER_PROFILE)

    # Weapons
    weapons = wb["Weapons"]
    weapon_headers = header_map(weapons)
    upsert_row_by_key(
        weapons,
        weapon_headers,
        "Weapon ID",
        CAVECRAWLER_WEAPON["Weapon ID"],
        CAVECRAWLER_WEAPON,
    )

    # Units
    units = wb["Units"]
    rotation_columns = [
        "Rotation Correction X Deg",
        "Rotation Correction Y Deg",
        "Rotation Correction Z Deg",
        "Turn Speed Deg/s",
    ]
    unit_headers = ensure_columns(units, ["Default Weapon ID"] + rotation_columns)
    stats_total = total_stats(CAVECRAWLER_UNIT)
    rating = power_rating(
        float(CAVECRAWLER_UNIT["Level"]),
        float(CAVECRAWLER_UNIT["Base HP"]),
        stats_total,
    )
    unit_row = dict(CAVECRAWLER_UNIT)
    unit_row["Rotation Correction Y Deg"] = 180
    unit_row["Turn Speed Deg/s"] = 360
    if "Total Stats" in unit_headers:
        unit_row["Total Stats"] = stats_total
    if "Power Rating" in unit_headers:
        unit_row["Power Rating"] = rating
    if "Tier" in unit_headers:
        unit_row["Tier"] = tier_label(rating)
    upsert_row_by_key(units, unit_headers, "Unit ID", "U-0003", unit_row)

    # Per-asset visual forward correction (UNIT-FACING-2) and turn speed (UNIT-TURN-1).
    rotation_corrections = {
        "U-0001": {"Rotation Correction Y Deg": 90, "Turn Speed Deg/s": 540},
        "U-0002": {"Turn Speed Deg/s": 720},
        "U-0003": {"Rotation Correction Y Deg": 180, "Turn Speed Deg/s": 360},
    }
    unit_id_col = unit_headers["Unit ID"]
    for row in range(2, units.max_row + 1):
        unit_id = str(units.cell(row, unit_id_col).value or "").strip()
        if unit_id not in rotation_corrections:
            continue
        for column, value in rotation_corrections[unit_id].items():
            units.cell(row, unit_headers[column], value)

    # Blank Default Weapon ID fails import when the column exists; fill legacy rows.
    weapon_col = unit_headers["Default Weapon ID"]
    defaults = {
        "U-0001": "weapon_fists",
        "U-0002": "weapon_claws",
    }
    for unit_id, weapon_id in defaults.items():
        col = unit_headers["Unit ID"]
        for row in range(2, units.max_row + 1):
            if str(units.cell(row, col).value or "").strip() == unit_id:
                if units.cell(row, weapon_col).value in (None, ""):
                    units.cell(row, weapon_col, weapon_id)
                break

    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")
    print(f"Backup: {BACKUP}")


if __name__ == "__main__":
    main()
