#!/usr/bin/env py -3
"""Update Chasma Design.xlsx Units sheet to match unit import schema."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Chasma Design.xlsx"
BACKUP = ROOT / "Chasma Design.xlsx.bak"

NEW_COLUMNS = [
    ("Move Speed", "USED — meters/sec for movement simulation (authoritative for speed)"),
    ("Collision Radius", "USED — meters; selection ring + steering radius"),
    ("Max Slope", "USED — degrees; max walkable terrain slope for pathfinding"),
    ("Render Scale", "USED — uniform glTF scale (1.0 = as authored; robot ≈ 2.15 for human height)"),
    ("Enabled", "USED — Y/N; N excludes row from imported catalog"),
]

HEADER_NOTES = {
    "Unit ID": "USED — unique id (UnitDefinitionId); required",
    "Name": "USED — display name; required",
    "Faction": "STORED — design metadata only; not runtime ownership",
    "File Path": "USED — GLB path → render key (e.g. robot); blank = no mesh",
    "Level": "STORED — catalog stat; no gameplay effect yet",
    "Base HP": "STORED — catalog stat; no gameplay effect yet",
    "Strength": "STORED — catalog stat; no gameplay effect yet",
    "Dexterity": "STORED — catalog stat; no gameplay effect yet",
    "Constitution": "STORED — catalog stat; no gameplay effect yet",
    "Agility": "STORED — catalog stat only; does NOT affect move speed (use Move Speed)",
    "Charisma": "STORED — catalog stat; no gameplay effect yet",
    "Intelligence": "STORED — catalog stat; no gameplay effect yet",
    "Total Stats": "IGNORED — sheet formula only; importer skips this column",
    "Power Rating": "STORED — imported float; no gameplay effect yet (script writes computed value)",
    "Tier": "USED — required string (Elite/Veteran/…); imported to catalog (script writes computed value)",
}

SCHEMA_ROWS = [
    ("Column", "Import", "Runtime use", "Notes"),
    ("Unit ID", "Required", "Catalog key", ""),
    ("Name", "Required", "Display name", ""),
    ("Faction", "Required", "Metadata tag", "Not instance ownership"),
    ("File Path", "Optional", "Render mesh key", "Blank = no dev spawn mesh"),
    ("Level", "Required", "Catalog only", ""),
    ("Base HP", "Required", "Catalog only", ""),
    ("Strength", "Required", "Catalog only", ""),
    ("Dexterity", "Required", "Catalog only", ""),
    ("Constitution", "Required", "Catalog only", ""),
    ("Agility", "Required", "Catalog only", "Does not drive movement speed"),
    ("Charisma", "Required", "Catalog only", ""),
    ("Intelligence", "Required", "Catalog only", ""),
    ("Total Stats", "Ignored", "—", "Excel helper column"),
    ("Power Rating", "Required", "Catalog only", ""),
    ("Tier", "Required", "Catalog only", ""),
    ("Move Speed", "Optional (default 4.0)", "Movement sim", "m/s — authoritative"),
    ("Collision Radius", "Optional (default 0.5)", "Movement + UI", "meters"),
    ("Max Slope", "Optional (default 40)", "Pathfinding", "degrees"),
    ("Render Scale", "Optional (default 1.0)", "Render mesh", "Uniform glTF scale"),
    ("Enabled", "Optional (default Y)", "Catalog filter", "N = row skipped"),
]


def total_stats(row_stats: list[float]) -> float:
    return sum(row_stats)


def power_rating(level: float, base_hp: float, stats: list[float]) -> float:
    return round(level * 2 + base_hp * 0.5 + total_stats(stats) * 0.8, 1)


def tier_label(rating: float) -> str:
    if rating >= 25:
        return "Elite"
    if rating >= 15:
        return "Veteran"
    if rating >= 8:
        return "Regular"
    return "Rookie"


def refresh_computed_columns(ws, headers: dict[str, int], row: int) -> None:
    """Write import-safe values for formula columns (openpyxl drops formula cache on save)."""
    level = float(ws.cell(row, headers["Level"]).value or 0)
    base_hp = float(ws.cell(row, headers["Base HP"]).value or 0)
    stats = [
        float(ws.cell(row, headers[name]).value or 0)
        for name in (
            "Strength",
            "Dexterity",
            "Constitution",
            "Agility",
            "Charisma",
            "Intelligence",
        )
    ]
    total = total_stats(stats)
    rating = power_rating(level, base_hp, stats)
    ws.cell(row, headers["Total Stats"], total)
    ws.cell(row, headers["Power Rating"], rating)
    ws.cell(row, headers["Tier"], tier_label(rating))


def is_unit_data_row(ws, row: int) -> bool:
    unit_id = ws.cell(row, 1).value
    if unit_id is None:
        return False
    text = str(unit_id).strip()
    if not text.startswith("U-"):
        return False
    name = ws.cell(row, 2).value
    return name is not None and str(name).strip() != ""


def suggested_move_speed(agility: float, name: str) -> float:
    if name.strip().lower() == "robot":
        return 9.0
    # Readable sheet default; code reads this column only.
    return round(3.5 + float(agility) * 0.35, 1)


def suggested_collision(strength: float, constitution: float) -> float:
    base = 0.45 + (float(strength) + float(constitution)) * 0.012
    return round(min(max(base, 0.4), 1.2), 2)


def suggested_render_scale(name: str) -> float:
    if name.strip().lower() == "robot":
        return 2.15
    return 1.0


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    shutil.copy2(WORKBOOK, BACKUP)

    wb = openpyxl.load_workbook(WORKBOOK)
    if "Units" not in wb.sheetnames:
        raise SystemExit("Missing Units sheet")

    ws = wb["Units"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

    start_col = ws.max_column + 1
    for offset, (header, note) in enumerate(NEW_COLUMNS):
        col = start_col + offset
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="006100")
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.comment = Comment(note, "chasma-import")

    # Map new column indices
    col_move = headers.get("Move Speed") or start_col
    col_collision = headers.get("Collision Radius") or (start_col + 1)
    col_slope = headers.get("Max Slope") or (start_col + 2)
    col_render_scale = headers.get("Render Scale") or (start_col + 3)
    col_enabled = headers.get("Enabled") or (start_col + 4)

    agility_col = headers["Agility"]
    strength_col = headers["Strength"]
    constitution_col = headers["Constitution"]
    name_col = headers["Name"]

    for row in range(2, ws.max_row + 1):
        if not is_unit_data_row(ws, row):
            continue
        agility = ws.cell(row, agility_col).value or 5
        strength = ws.cell(row, strength_col).value or 5
        constitution = ws.cell(row, constitution_col).value or 5
        name = str(ws.cell(row, name_col).value or "")

        if ws.cell(row, col_move).value in (None, ""):
            ws.cell(row, col_move, suggested_move_speed(agility, name))
        if ws.cell(row, col_collision).value in (None, ""):
            ws.cell(row, col_collision, suggested_collision(strength, constitution))
        if ws.cell(row, col_slope).value in (None, ""):
            ws.cell(row, col_slope, 40)
        if ws.cell(row, col_render_scale).value in (None, ""):
            ws.cell(row, col_render_scale, suggested_render_scale(name))
        if ws.cell(row, col_enabled).value in (None, ""):
            ws.cell(row, col_enabled, "Y")

        refresh_computed_columns(ws, headers, row)

    # Header comments on existing columns
    for header, note in HEADER_NOTES.items():
        if header in headers:
            cell = ws.cell(1, headers[header])
            cell.comment = Comment(note, "chasma-import")

    # Schema reference sheet
    if "Unit Import Schema" in wb.sheetnames:
        del wb["Unit Import Schema"]
    schema = wb.create_sheet("Unit Import Schema", 0)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for r, row in enumerate(SCHEMA_ROWS, start=1):
        for c, value in enumerate(row, start=1):
            cell = schema.cell(r, c, value)
            if r == 1:
                cell.font = title_font
                cell.fill = header_fill
    schema.column_dimensions["A"].width = 18
    schema.column_dimensions["B"].width = 22
    schema.column_dimensions["C"].width = 22
    schema.column_dimensions["D"].width = 48
    schema.freeze_panes = "A2"

    note = schema.cell(
        len(SCHEMA_ROWS) + 2,
        1,
        "Importer: src/data_import/unit/schema.rs + excel.rs. "
        "Dev catalog loads this sheet at startup (feature=dev). "
        "Starter catalog (wolf/bandit/deer) is fallback only. "
        "Total Stats / Power Rating / Tier are written as plain values so calamine can read them "
        "(Excel formulas are not re-evaluated on import).",
    )
    note.alignment = Alignment(wrap_text=True)
    schema.merge_cells(
        start_row=len(SCHEMA_ROWS) + 2,
        start_column=1,
        end_row=len(SCHEMA_ROWS) + 2,
        end_column=4,
    )

    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")
    print(f"Backup: {BACKUP}")


if __name__ == "__main__":
    main()
