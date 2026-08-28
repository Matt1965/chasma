"""Add Sight Range column to the Units sheet (ADR-132 Phase 4)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"
SIGHT_RANGE_COLUMN = "Sight Range"
DEFAULT_SIGHT_RANGE = 24.0


def update_units(wb: openpyxl.Workbook) -> None:
    ws = wb["Units"]
    headers = [cell.value for cell in ws[1]]
    if SIGHT_RANGE_COLUMN in headers:
        col_idx = headers.index(SIGHT_RANGE_COLUMN)
    else:
        col_idx = len(headers)
        ws.cell(row=1, column=col_idx + 1, value=SIGHT_RANGE_COLUMN)

    for row in range(2, ws.max_row + 1):
        unit_id = ws.cell(row=row, column=headers.index("Unit ID") + 1).value
        if not unit_id or not isinstance(unit_id, str) or not unit_id.strip():
            continue
        cell = ws.cell(row=row, column=col_idx + 1)
        if cell.value in (None, ""):
            cell.value = DEFAULT_SIGHT_RANGE


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    update_units(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK} with {SIGHT_RANGE_COLUMN} defaults ({DEFAULT_SIGHT_RANGE} m)")


if __name__ == "__main__":
    main()
