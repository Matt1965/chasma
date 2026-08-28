"""Phase 2 workbook migration: relationship matrices + legacy Factions cleanup."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"

FACTION_KEYS = [
    "player",
    "wild",
    "bandits",
    "raiders",
    "nomads",
    "town",
    "slavers",
    "tech_hunters",
    "scavs",
    "dominion",
]

SPECIES_KEYS = ["robot", "fox", "cavecrawler"]


def cleanup_factions_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb["Factions"]
    headers = [cell.value for cell in ws[1]]
    if "Disposition" in headers:
        col_idx = headers.index("Disposition") + 1
        ws.delete_cols(col_idx)
        headers = [cell.value for cell in ws[1]]

    # Remove legacy footer / relationship table rows below authored factions.
    legacy_start = None
    for row in range(2, ws.max_row + 1):
        legacy = ws.cell(row=row, column=headers.index("Faction ID") + 1).value
        name = ws.cell(row=row, column=headers.index("Name") + 1).value
        if isinstance(name, str) and name.startswith("FACTION RELATIONSHIPS"):
            legacy_start = row
            break
        if legacy is None and isinstance(name, str) and name.strip():
            legacy_start = row
            break
        if legacy is not None and not (isinstance(legacy, str) and legacy.startswith("F-")):
            legacy_start = row
            break

    if legacy_start is not None:
        ws.delete_rows(legacy_start, ws.max_row - legacy_start + 1)


def write_rel_faction_faction(wb: openpyxl.Workbook) -> None:
    name = "Rel Faction Faction"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value="Faction -> Faction")
    for col, key in enumerate(FACTION_KEYS, start=2):
        ws.cell(row=1, column=col, value=key)
    for row, key in enumerate(FACTION_KEYS, start=2):
        ws.cell(row=row, column=1, value=key)


def write_rel_faction_species(wb: openpyxl.Workbook) -> None:
    name = "Rel Faction Species"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value="Faction -> Species")
    for col, key in enumerate(SPECIES_KEYS, start=2):
        ws.cell(row=1, column=col, value=key)
    for row, key in enumerate(["player", "wild"], start=2):
        ws.cell(row=row, column=1, value=key)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    cleanup_factions_sheet(wb)
    write_rel_faction_faction(wb)
    write_rel_faction_species(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")


if __name__ == "__main__":
    main()
