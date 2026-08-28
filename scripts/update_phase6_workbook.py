"""Phase 6 workbook migration: author wild -> player relationship edge."""

from __future__ import annotations

from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"
SHEET = "Rel Faction Faction"
WILD_TO_PLAYER = -300


def set_wild_to_player_edge(wb: openpyxl.Workbook) -> None:
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"missing worksheet `{SHEET}`")
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    if not headers or headers[0] != "Faction -> Faction":
        raise RuntimeError(f"unexpected `{SHEET}` header row")

    col_player = headers.index("player") + 1
    row_wild = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "wild":
            row_wild = row
            break
    if row_wild is None:
        raise RuntimeError("missing `wild` row in Rel Faction Faction")

    ws.cell(row=row_wild, column=col_player, value=WILD_TO_PLAYER)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    set_wild_to_player_edge(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}: wild -> player = {WILD_TO_PLAYER}")


if __name__ == "__main__":
    main()
