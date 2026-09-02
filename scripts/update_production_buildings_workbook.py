"""Add stone_quarry and prispod_farm production buildings to the Buildings sheet."""

from __future__ import annotations

from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"

BUILDINGS = [
    {
        "Building ID": "stone_quarry",
        "Name": "Stone Quarry",
        "Category": "production",
        "Model File Path": "assets/buildings/stone_mine.glb",
        "Collision File Path": "assets/buildings/stone_mine.glb",
        "Health": 450,
        "Build Time": 100,
        "Footprint Type": "Rectangle",
        "Footprint Width": 16,
        "Footprint Depth": 12,
        "Max Slope": 30,
        "Enabled": "Y",
    },
    {
        "Building ID": "prispod_farm",
        "Name": "Prispod Farm",
        "Category": "production",
        "Model File Path": "assets/buildings/prispod_farm.glb",
        "Collision File Path": "assets/buildings/prispod_farm.glb",
        "Health": 300,
        "Build Time": 80,
        "Footprint Type": "Rectangle",
        "Footprint Width": 26,
        "Footprint Depth": 20,
        "Max Slope": 35,
        "Enabled": "Y",
    },
]


def upsert_building_row(ws, headers: list[str | None], row_values: dict[str, object]) -> None:
    id_col = headers.index("Building ID")
    building_id = row_values["Building ID"]
    target_row = None
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=id_col + 1).value
        if cell and str(cell).strip() == building_id:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1

    for header, value in row_values.items():
        if header not in headers:
            continue
        col = headers.index(header) + 1
        ws.cell(row=target_row, column=col, value=value)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Buildings"]
    headers = [cell.value for cell in ws[1]]
    for row_values in BUILDINGS:
        upsert_building_row(ws, headers, row_values)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK} with production buildings: stone_quarry, prispod_farm")


if __name__ == "__main__":
    main()
