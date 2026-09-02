"""Settlement AI Phase 4 workbook updates: nutrition, prispod, construction_material."""

from __future__ import annotations

from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"

NUTRITION_COLUMN = "Nutrition"
FOOD_NUTRITION = {
    "dried_meat": 20,
    "cactus_flesh": 10,
    "prispod": 25,
}


def ensure_nutrition_column(ws, headers: list[str | None]) -> int:
    if NUTRITION_COLUMN in headers:
        return headers.index(NUTRITION_COLUMN)
    col_idx = len(headers)
    ws.cell(row=1, column=col_idx + 1, value=NUTRITION_COLUMN)
    headers.append(NUTRITION_COLUMN)
    return col_idx


def update_item_categories(wb: openpyxl.Workbook) -> None:
    ws = wb["Item Categories"]
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("Category ID")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col + 1).value == "construction_material":
            return
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=headers.index("Category ID") + 1, value="construction_material")
    ws.cell(row=next_row, column=headers.index("Name") + 1, value="Construction Material")
    ws.cell(row=next_row, column=headers.index("Enabled") + 1, value="Y")
    if "Description" in headers:
        ws.cell(
            row=next_row,
            column=headers.index("Description") + 1,
            value="Building and construction resources",
        )
    if "Sort Priority" in headers:
        ws.cell(row=next_row, column=headers.index("Sort Priority") + 1, value=15)


def update_items(wb: openpyxl.Workbook) -> None:
    ws = wb["Items"]
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("Item ID")
    category_col = headers.index("Category")
    nutrition_col = ensure_nutrition_column(ws, headers)

    prispod_exists = False
    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=id_col + 1).value
        if not item_id or not isinstance(item_id, str) or not item_id.strip():
            continue
        item_id = item_id.strip()
        if item_id == "stone":
            ws.cell(row=row, column=category_col + 1, value="construction_material")
        if item_id in FOOD_NUTRITION:
            ws.cell(row=row, column=nutrition_col + 1, value=FOOD_NUTRITION[item_id])
        if item_id == "prispod":
            prispod_exists = True

    if not prispod_exists:
        next_row = ws.max_row + 1
        values = {
            "Item ID": "prispod",
            "Name": "Prispod",
            "Category": "food",
            "Width": 1,
            "Height": 1,
            "Stackable": "Y",
            "Max Stack": 20,
            "Mass Grams": 250,
            "Enabled": "Y",
            "Description": "Edible prispod cultivated on farms.",
            "Base Value": 5,
            NUTRITION_COLUMN: 25,
        }
        for header, value in values.items():
            if header in headers:
                ws.cell(row=next_row, column=headers.index(header) + 1, value=value)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    update_item_categories(wb)
    update_items(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK} for Settlement AI Phase 4 content")


if __name__ == "__main__":
    main()
