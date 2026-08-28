"""One-shot Phase 1 workbook migration for relationship identity sheets."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "Chasma Design.xlsx"

FACTION_SLUGS = {
    "Player": "player",
    "Wild": "wild",
    "Bandits": "bandits",
    "Raiders": "raiders",
    "Nomads": "nomads",
    "Town": "town",
    "Slavers": "slavers",
    "Tech Hunters": "tech_hunters",
    "Scavs": "scavs",
    "The Dominion": "dominion",
}

UNIT_SPECIES = {
    "Robot": ("player", "robot"),
    "Fox": ("player", "fox"),
    "Cavecrawler": ("wild", "cavecrawler"),
}

SPECIES_ROWS = [
    ("robot", "Robot", "Y", "Player humanoid construct"),
    ("fox", "Fox", "Y", "Small canid companion"),
    ("cavecrawler", "Cavecrawler", "Y", "Subterranean predator"),
]


def slugify(name: str) -> str:
    if name in FACTION_SLUGS:
        return FACTION_SLUGS[name]
    key = name.strip().lower()
    key = re.sub(r"^the\s+", "", key)
    key = re.sub(r"[^a-z0-9]+", "_", key)
    return key.strip("_")


def update_factions(wb: openpyxl.Workbook) -> None:
    ws = wb["Factions"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    if "Faction Key" not in col:
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value="Faction Key")
        headers = [cell.value for cell in ws[1]]
        col = {name: idx for idx, name in enumerate(headers) if name}

    if "Enabled" not in col:
        insert_at = col.get("Description", len(headers) - 1) + 2
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value="Enabled")
        headers = [cell.value for cell in ws[1]]
        col = {name: idx for idx, name in enumerate(headers) if name}

    name_idx = col["Name"]
    legacy_idx = col.get("Faction ID")
    key_idx = col["Faction Key"]
    enabled_idx = col["Enabled"]

    for row in range(2, ws.max_row + 1):
        legacy = (
            ws.cell(row=row, column=legacy_idx + 1).value
            if legacy_idx is not None
            else None
        )
        name = ws.cell(row=row, column=name_idx + 1).value
        if not name or not isinstance(name, str):
            continue
        if legacy_idx is not None and not (isinstance(legacy, str) and legacy.startswith("F-")):
            continue
        slug = slugify(name)
        ws.cell(row=row, column=key_idx + 1, value=slug)
        if not ws.cell(row=row, column=enabled_idx + 1).value:
            ws.cell(row=row, column=enabled_idx + 1, value="Y")


def update_species(wb: openpyxl.Workbook) -> None:
    if "Species" in wb.sheetnames:
        del wb["Species"]
    ws = wb.create_sheet("Species")
    ws.append(["Species Key", "Name", "Enabled", "Description"])
    for row in SPECIES_ROWS:
        ws.append(list(row))


def update_units(wb: openpyxl.Workbook) -> None:
    ws = wb["Units"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    if "Faction" in col and "Faction Key" not in col:
        faction_idx = col["Faction"]
        ws.cell(row=1, column=faction_idx + 1, value="Faction Key")
        headers[faction_idx] = "Faction Key"
        col["Faction Key"] = faction_idx
        del col["Faction"]

    if "Species Key" not in col:
        insert_at = col["Faction Key"] + 2
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value="Species Key")
        headers = [cell.value for cell in ws[1]]
        col = {name: idx for idx, name in enumerate(headers) if name}

    name_idx = col["Name"]
    faction_idx = col["Faction Key"]
    species_idx = col["Species Key"]

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_idx + 1).value
        if not name:
            continue
        if name in UNIT_SPECIES:
            faction_key, species_key = UNIT_SPECIES[name]
        else:
            faction_cell = ws.cell(row=row, column=faction_idx + 1).value
            faction_key = slugify(str(faction_cell)) if faction_cell else "wild"
            species_key = slugify(str(name))
        ws.cell(row=row, column=faction_idx + 1, value=faction_key)
        ws.cell(row=row, column=species_idx + 1, value=species_key)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    update_factions(wb)
    update_species(wb)
    update_units(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")


if __name__ == "__main__":
    main()
